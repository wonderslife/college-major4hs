from collections import defaultdict
import openpyxl

# 读取教育部第四次全国高校学科评估结果文件
def read_4th_evaluation(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        school_name = row[6]  # 院校名称在第7列
        major_name = row[3]  # 一级学科名称在第3列
        assessment_result = row[5]  # 评估结果在第6列
        major_code = row[4] # 学科代码
        data.append((school_name, major_name, assessment_result, major_code))
    wb.close()
    return data

# 读取5th_evaluation文件
def read_5th_evaluation(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    rating_dict = defaultdict(str)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        school_name = row[0]
        major_name = row[1]
        rating = row[2]
        rating_dict[(school_name, major_name)] = rating
    wb.close()
    return rating_dict

# 使用字典查找来替换评估结果
def replace_assessment_result(data, rating_dict):
    batch = ""
    result = []
    for school_name, major_name, assessment_result, major_code in data:
        key = (school_name, major_name)
        if key in rating_dict:
            new_assessment_result = rating_dict[key]
            batch = "5th"
        else:
            new_assessment_result = assessment_result
            batch = "4th"
        result.append((school_name, major_code, major_name, new_assessment_result, batch))
    return result

# 保存修改后的结果
def save_result(result, file_path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['院校名称', '学科代码','一级学科名称', '评估结果', '评估批次'])
    for school_name, major_code, major_name, assessment_result, batch in result:
        sheet.append([school_name, major_code, major_name, assessment_result, batch])
    wb.save(file_path)
    wb.close()

if __name__ == '__main__':
    # 读取教育部第四次全国高校学科评估结果文件
    fourth_evaluation_data = read_4th_evaluation('教育部\\第四次学科评估结果\\教育部第四次全国高校学科评估结果.xlsx')

    # 读取5th_evaluation文件
    fiveth_rating_dict = read_5th_evaluation('教育部\\5th\\5th_evaluation.xlsx')

    # 使用字典查找来替换评估结果
    replaced_data = replace_assessment_result(fourth_evaluation_data, fiveth_rating_dict)

    # 保存修改后的结果
    save_result(replaced_data, '教育部\\4.5_evaluation.xlsx')