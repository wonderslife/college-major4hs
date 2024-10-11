import json
import logging
import os
import time
import openpyxl
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from multiprocessing import Pool

logging.basicConfig(level=logging.INFO)

def extract_qa_data(driver, name):
    qalist = []
    try:
        more_link = driver.find_element(By.XPATH, "/html/body/div[1]/div[4]/div/div[2]/div[2]/div[1]/a").get_attribute("href")
    except NoSuchElementException:
        return qalist
    logging.info(f'进程:{os.getpid()} 正在解析 {name} QALink: {more_link}')
    driver.get(more_link)
    time.sleep(1)
    # 获取总页数，分页页面的倒数第二个
    try:
        page_content = driver.find_element(By.CSS_SELECTOR, "ul.ch-page.clearfix").find_elements(By.TAG_NAME, "li")
    except NoSuchElementException:
        return qalist
    if len(page_content) >= 2:
        pagecount = int(page_content[-2].text)
    # 全部展开
    driver.find_element(By.CSS_SELECTOR, "a#allShow").click()
    time.sleep(1)
    qa_contents = driver.find_elements(By.CSS_SELECTOR, "div.qa-box.qa-box-other")
    for qa in qa_contents:
        try:
            title = qa.find_element(By.CSS_SELECTOR, "a.q-name-link").text
            question = qa.find_element(By.CSS_SELECTOR, "div.a-title.UEditor").text
            answer = qa.find_element(By.CSS_SELECTOR, "div.a-content-box.UEditor").find_element(By.TAG_NAME, "p").text
            qalist.append({"schoolName": name, "title": name + "-" + title, "question": question, "answer": answer})
        except NoSuchElementException:
            continue
    for page in range(2, pagecount + 1):
        logging.info(f'进程:{os.getpid()} 正在解析 {name} 第 {page} 页')
        if page > 1:
            try:
                # 点击 > 按钮,分页中的最后一个
                next_page_button = driver.find_element(By.LINK_TEXT, ">")
                next_page_button.click()
                time.sleep(1)
            except NoSuchElementException:
                break
        driver.find_element(By.CSS_SELECTOR, "a#allShow").click()
        time.sleep(1)
        qa_contents = driver.find_elements(By.CSS_SELECTOR, "div.qa-box.qa-box-other")
        for qa in qa_contents:
            try:
                title = qa.find_element(By.CSS_SELECTOR, "a.q-name-link").text
                question = qa.find_element(By.CSS_SELECTOR, "div.a-title.UEditor").text
                answer = qa.find_element(By.CSS_SELECTOR, "div.a-content-box.UEditor").find_element(By.TAG_NAME, "p").text
                qalist.append({"schoolName": name, "title": name + "-" + title, "question": question, "answer": answer})
            except NoSuchElementException:
                continue
    return qalist

def send_email_notification(subject, message):
    sender_email = "qikun@163.com"
    receiver_email = "qikun@163.com"
    password = "MScGJCa2fXbNL8aC"
    msg = MIMEText(message)
    msg['Subject'] = subject
    msg['From'] = sender_email
    msg['To'] = receiver_email
    try:
        with smtplib.SMTP_SSL('smtp.163.com', 465) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, msg.as_string())
    except Exception as e:
        logging.warning(f'进程:{os.getpid()} 邮件发送失败:{e},继续执行程序。')

def process_data(start_row, end_row, file_name):
    driver = webdriver.Chrome()
    #wb = openpyxl.load_workbook('教育部\\yangguanggaokao_20241003.xlsx')
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    local_file_count = 0
    combined_qalist = []
    for index, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True), start=start_row):
        sclink = row[5]
        name = row[0]
        if sclink:
            logging.info(f'进程:{os.getpid()} {name} 链接地址:{sclink}')
            driver.get(sclink)
            time.sleep(1)
            qalist = extract_qa_data(driver, name)
            combined_qalist.extend(qalist)
            if index % 100 == 0:
                current_date = datetime.now().strftime('%Y%m%d%H%M%S')
                with open(f'C:\\Users\\wonder\\Documents\\college-major4hs\\教育部\\qa\\qa-{current_date}-{local_file_count}.json', 'w', encoding='utf-8') as json_file:
                    json.dump(combined_qalist, json_file, ensure_ascii=False, indent=4)
                logging.info(f'进程:{os.getpid()} 保存阶段文件{local_file_count}')
                # 发送邮件通知
                subject = f'文件保存通知 - qa-{current_date}-{local_file_count}.json'
                message = f'进程:{os.getpid()} 已保存过程文件：qa-{current_date}-{local_file_count}.json'
                send_email_notification(subject, message)
                local_file_count += 1
                combined_qalist = []
    driver.quit()
    return combined_qalist

if __name__ == '__main__':
    #file_name = '教育部\yangguanggaokao_20241003.xlsx'
    file_name = '教育部\\qa_missing_schools.xlsx'
    wb = openpyxl.load_workbook(file_name)
    
    ws = wb.active
    total_rows = ws.max_row - 1
    batch_size = 200
    processes = 3
    pool = Pool(processes)
    results = []
    for i in range(0, total_rows, batch_size):
        start_row = i + 2
        end_row = min(i + batch_size + 1, total_rows + 2)
        results.append(pool.apply_async(process_data, args=(start_row, end_row, file_name)))
    pool.close()
    pool.join()
    combined_qalist = []
    for result in results:
        combined_qalist.extend(result.get())
    # 保存剩余数据
    if combined_qalist:
        current_date = datetime.now().strftime('%Y%m%d%H%M%S')
        with open(f'C:\\Users\\wonder\\Documents\\college-major4hs\\教育部\\qa\\qa-{current_date}-last.json', 'w', encoding='utf-8') as json_file:
            json.dump(combined_qalist, json_file, ensure_ascii=False, indent=4)
        # 发送邮件通知
        subject = f'最后文件保存通知 - qa-{current_date}-last.json'
        message = f'已保存最后文件：qa-{current_date}-last.json'
        send_email_notification(subject, message)