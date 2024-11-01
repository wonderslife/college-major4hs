import openpyxl
import logging
## 将位次填到录取表分数表中
# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 加载文件
score_file = "2024高考\\2024一分一段表.xlsx"
rank_file = "2024高考\\2024专业投档线位次.xlsx"

# 打开工作簿
score_wb = openpyxl.load_workbook(score_file)
rank_wb = openpyxl.load_workbook(rank_file)

# 选择工作表
score_sheet = score_wb.active
rank_sheet = rank_wb.active

# 创建分数到位次的映射
score_to_rank = {}
for row in score_sheet.iter_rows(min_row=2, values_only=True):
    score, personcount,rank = str(row[0]),str(row[1]),str(row[2])  # 第一列是分数，第二列是人数,第三列是位次
    score_to_rank[score] = rank
    logging.info(f'加载分数: {score}, 位次: {rank}')

# 填充位次到专业投档线位次表
for row in rank_sheet.iter_rows(min_row=2):
    score_cell = row[4]  
    if str(score_cell.value)in score_to_rank:
        rank_cell = row[5]  # 最后一列是要填充位次的列
        rank_cell.value = score_to_rank[str(score_cell.value)]
        logging.info(f'填充分数: {score_cell.value}, 位次: {score_to_rank[str(score_cell.value)]}')
    else:
        logging.warning(f'未找到分数: {score_cell.value}')

# 保存修改后的文件
rank_wb.save(rank_file)
logging.info('文件已保存: {}'.format(rank_file))
