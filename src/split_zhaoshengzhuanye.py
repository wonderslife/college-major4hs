import logging
from openpyxl import load_workbook

# 设置日志配置
logging.basicConfig(level=logging.INFO)

try:
# 加载 Excel 文件
    workbook = load_workbook('2024高考\\2024招生专业.xlsx')
    original_sheet = workbook.active

    # 获取最大行数
    max_row = original_sheet.max_row

    # 用于存储新的数据
    new_data = []

    # 遍历每一行，从第二行开始
    for row in range(2, max_row + 1):
        cell1 = original_sheet.cell(row=row, column=1).value
        cell2 = original_sheet.cell(row=row, column=2).value
        cell3 = original_sheet.cell(row=row, column=3).value
        cell4 = original_sheet.cell(row=row, column=4).value
        cell5 = original_sheet.cell(row=row, column=5).value
        cell6 = original_sheet.cell(row=row, column=6).value
        cell7 = original_sheet.cell(row=row, column=7).value
        cell8 = original_sheet.cell(row=row, column=8).value
        new_data.append((cell1, cell2, cell3, cell4, cell5, cell6, cell7, cell8))
        if cell6 and ';' in cell6:
            # 拆分专业
            professions = cell6.split(';')
            for prof in professions:
                prof_parts = prof.split('-')
                if len(prof_parts) == 2:
                    new_data.append((cell1, cell2, cell3, prof_parts[0].strip(), prof_parts[1].strip(), "", cell7, cell8))
        logging.info(f"处理到第 {row} 行")

    # 创建新的工作表
    new_sheet = workbook.create_sheet(title="Sheet1")

    # 将新数据写入新工作表
    row_index = 1
    for cell1, cell2, cell3, cell4, cell5, cell6, cell7, cell8 in new_data:
        new_sheet.cell(row=row_index, column=1, value=cell1)
        new_sheet.cell(row=row_index, column=2, value=cell2)
        new_sheet.cell(row=row_index, column=3, value=cell3)
        new_sheet.cell(row=row_index, column=4, value=cell4)
        new_sheet.cell(row=row_index, column=5, value=cell5)
        new_sheet.cell(row=row_index, column=6, value=cell6)
        new_sheet.cell(row=row_index, column=7, value=cell7)
        new_sheet.cell(row=row_index, column=8, value=cell8)
        row_index += 1
        logging.info(f"写入新工作表第 {row_index} 行数据")

    # 保存修改后的 Excel 文件

    workbook.save('2024高考\\2024招生专业评估.xlsx')
except Exception as e:
    logging.error(f"操作出错: {e}")