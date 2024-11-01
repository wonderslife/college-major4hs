import openpyxl
import logging

# 配置日志记录
logging.basicConfig(filename='operation_log.log', level=logging.INFO, force=True)

def match_and_update():
    # 加载 4.5_evaluation.xlsx 文件
    wb1 = openpyxl.load_workbook('教育部\\4.5_evaluation.xlsx')
    sheet1 = wb1.active

    # 加载 2024 招生专业评估.xlsx 文件
    wb2 = openpyxl.load_workbook('2024高考\\2024招生专业评估.xlsx')
    sheet2 = wb2.active

    try:
        # 创建查找字典，键为(学校名称, 专业代码)的元组，来自 wb1
        lookup_dict = {}
        for row1 in sheet1.iter_rows(min_row=2, values_only=True):
            key = (row1[0], row1[1])  # (学校名称, 专业代码)
            lookup_dict[key] = row1[3]  # 存储评估结果

        # 遍历专业评估文件并直接查找匹配项
        for row2 in sheet2.iter_rows(min_row=2):
            school_name = row2[2].value
            major_code = row2[3].value[:4]  # 取专业代码的前四位
            
            key = (school_name, major_code)
            if key in lookup_dict:
                eval_result = lookup_dict[key]
                row_num = row2[0].row
                sheet2.cell(row=row_num, column=9, value=eval_result)
                logging.info(f"已将评估结果 '{eval_result}' 写入行 {row_num} {school_name} {major_code}")

        wb2.save('2024高考\\2024招生专业评估.xlsx')
        logging.info("更新文件保存成功")
    except Exception as e:
        logging.error(f"操作出错: {e} row1:{row1} row2:{row2}")

if __name__ == '__main__':
    match_and_update()