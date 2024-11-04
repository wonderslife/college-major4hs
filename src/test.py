# 处理招生专业，以及专业录取，在招生专业中加入录取的分数和位次，增加列是录取专业名称和相似度
# 根据招生专业循环投档线，分数和位次，查看招生专业中没有录取的，也查看录取的但是没在招生专业中的
from openpyxl import load_workbook
import re
import difflib

# 读取Excel文件
wb = load_workbook('2024高考\\2024专业投档线位次.xlsx')
ws = wb.active

# 创建一个字典来存储专业分数和位次的映射关系
score_rank_map = {}

# 获取表头行
header_row = next(ws.rows)
headers = [cell.value for cell in header_row]

# 获取列索引
school_idx = headers.index('学校名称')
major_idx = headers.index('专业名称') 
score_idx = headers.index('投档最低分')
rank_idx = headers.index('位次')

# 遍历工作表的每一行(跳过表头)
for row in list(ws.rows)[1:]:
    school = row[school_idx].value
    major = row[major_idx].value
    score = row[score_idx].value
    rank = row[rank_idx].value
    
    # 如果学校不在字典中,创建一个新的列表
    if school not in score_rank_map:
        score_rank_map[school] = []
        
    # 将专业信息添加到对应学校的列表中
    major_info = {}
    major_info = {major: {'专业': major, '分数': score, '位次': rank}}
    score_rank_map[school].append(major_info)


########################################################
# 读取招生专业比对文件
wb_compare = load_workbook('2024高考\\2024招生专业-比对.xlsx')
ws_compare = wb_compare.active

# 获取比对文件的表头
compare_headers = [cell.value for cell in next(ws_compare.rows)]

# 获取列索引
compare_school_idx = compare_headers.index('学校名称')
compare_major_idx = compare_headers.index('专业名称')
compare_submajor_idx = compare_headers.index('包含专业代码及名称')
compare_rank_idx = compare_headers.index('位次')
compare_score_idx = compare_headers.index('分数')
compare_admit_major_idx = compare_headers.index('录取专业')
compare_similarity_idx = compare_headers.index('相似度')

# 遍历比对文件的每一行
for row in list(ws_compare.rows)[1:]:
    school = row[compare_school_idx].value
    major = row[compare_major_idx].value
    submajor = row[compare_submajor_idx].value
    
    # 在score_rank_map中查找相同学校的记录
    if submajor:
        # 去掉数字和-字符,生成新的专业名称
        cleaned_submajor = re.sub(r'\d+-', '', submajor)
        new_major = f"{major}({cleaned_submajor})"
        
        # 在score_rank_map中查找该学校
        if school in score_rank_map:
            # 遍历该学校的所有专业
            for major_info in score_rank_map[school]:
                
                for stored_major, details in major_info.items():
                    # 计算专业名称相似度
                    similarity = difflib.SequenceMatcher(None, stored_major, new_major).ratio()
                    print(f"相似度:{similarity} {stored_major} {new_major}")
                    # 如果相似度大于0.95,认为是相同专业
                    if similarity > 0.95:
                        # 更新excel中的位次信息
                        row[compare_rank_idx].value = details['位次']
                        row[compare_score_idx].value = details['分数']
                        row[compare_admit_major_idx].value = stored_major
                        row[compare_similarity_idx].value = similarity

    else:
    # 在score_rank_map中查找该学校
        if school in score_rank_map:
            # 遍历该学校的所有专业
            for major_info in score_rank_map[school]:
                for stored_major, details in major_info.items():
                    # 计算专业名称相似度
                    similarity = difflib.SequenceMatcher(None, stored_major, major).ratio()
                    print(f"相似度:{similarity} {stored_major} {major}")
                    # 如果相似度大于0.95,认为是相同专业
                    if similarity > 0.95:
                        # 更新excel中的位次信息
                        row[compare_rank_idx].value = details['位次']
                        row[compare_score_idx].value = details['分数']
                        row[compare_admit_major_idx].value = stored_major
                        row[compare_similarity_idx].value = similarity


# 保存修改后的文件
wb_compare.save('2024高考\\2024招生专业-比对-更新.xlsx')

